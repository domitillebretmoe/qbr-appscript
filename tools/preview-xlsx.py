#!/usr/bin/env python3
"""Turn a sheet recorded by test/render.test.js --dump into an .xlsx preview of what Render.gs draws.

    node test/render.test.js --dump /tmp/preview.json Q2-2026 && python3 tools/preview-xlsx.py /tmp/preview.json preview.xlsx

Google-only features are approximated: SPARKLINE formulas become unicode sparklines, the percentage
gradient becomes an Excel 2-colour scale, and charts are rebuilt as Excel charts from the same data.
"""
import json
import sys

from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.filters import AutoFilter

SPARK = '\u2581\u2582\u2583\u2584\u2585\u2586\u2587\u2588'


def color(hex_code):
    return hex_code.lstrip('#').upper() if hex_code and hex_code != 'transparent' else None


def spark(values):
    nums = [v for v in values if isinstance(v, (int, float))]
    if len(nums) < 2 or min(nums) == max(nums):
        return SPARK[3] * max(1, len(nums))
    lo, hi = min(nums), max(nums)
    return ''.join(SPARK[int((v - lo) / (hi - lo) * 7)] for v in nums)


def main(src, dst):
    dumps = json.load(open(src))
    wb = Workbook()
    for i, data in enumerate(dumps if isinstance(dumps, list) else [dumps]):
        ws = wb.active if i == 0 else wb.create_sheet()
        write_sheet(ws, data)
    wb.save(dst)


def write_sheet(ws, data):
    cells = data['cells']
    ws.title = data['name']
    ws.sheet_view.showGridLines = not data['hiddenGridlines']
    if data['frozenRows'] or data.get('frozenColumns'):
        ws.freeze_panes = f'{get_column_letter(data.get("frozenColumns", 0) + 1)}{data["frozenRows"] + 1}'
    for col, px in data['colWidths'].items():
        ws.column_dimensions[get_column_letter(int(col))].width = px / 7.2
    for row, px in data['rowHeights'].items():
        ws.row_dimensions[int(row)].height = px * 0.75

    def range_values(a1):
        c1, r1, c2, r2 = range_boundaries(a1)
        return [cells.get(f'{r},{c}', {}).get('value') for r in range(r1, r2 + 1) for c in range(c1, c2 + 1)]

    shown = {}
    for key, cell in cells.items():
        c = ws.cell(row=cell['row'], column=cell['col'])
        value = cell.get('value')
        if isinstance(value, str) and value.startswith('=SPARKLINE('):
            value = spark(range_values(value[len('=SPARKLINE('):value.index(',')]))
        shown[key] = value
        if value not in (None, ''):
            c.value = value
        if cell.get('numberFormat') and cell.get('numberFormat') != '@':
            c.number_format = cell['numberFormat']
        font_name = cell.get('fontFamily', 'Arial')
        c.font = Font(
            name=font_name, size=cell.get('fontSize', 10), bold=cell.get('fontWeight') == 'bold',
            italic=cell.get('fontStyle') == 'italic', color=color(cell.get('fontColor')),
        )
        if cell.get('background'):
            c.fill = PatternFill('solid', start_color=color(cell['background']))
        c.alignment = Alignment(
            horizontal=cell.get('hAlign'), vertical={'middle': 'center'}.get(cell.get('vAlign'), cell.get('vAlign')),
            wrap_text=bool(cell.get('wrap')),
        )
        if cell.get('richText'):
            text = str(value)
            c.value = CellRichText(*[
                TextBlock(InlineFont(rFont=font_name, sz=run['style'].get('fontSize', 10), b=run['style'].get('bold', False),
                                     color=color(run['style'].get('color'))), text[run['start']:run['end']])
                for run in cell['richText']
            ])

    for a1 in data['merges']:
        ws.merge_cells(a1)

    for b in data['borders']:
        side = Side(style='medium' if b.get('style') == 'SOLID_MEDIUM' else 'thin', color=color(b['color']))
        c1, r1, c2, r2 = range_boundaries(b['range'])
        for r in range(r1, r2 + 1):
            for cc in range(c1, c2 + 1):
                cell = ws.cell(row=r, column=cc)
                cur = cell.border
                cell.border = Border(
                    top=side if (b['top'] and r == r1) or (b['horizontal'] and r > r1) else cur.top,
                    bottom=side if (b['bottom'] and r == r2) or (b['horizontal'] and r < r2) else cur.bottom,
                    left=side if (b['left'] and cc == c1) or (b['vertical'] and cc > c1) else cur.left,
                    right=side if (b['right'] and cc == c2) or (b['vertical'] and cc < c2) else cur.right,
                )

    for rule in data['rules']:
        for ref in rule.get('ranges', []):
            ws.conditional_formatting.add(ref, ColorScaleRule(
                start_type='num', start_value=float(rule['min']['value']), start_color=color(rule['min']['color']),
                end_type='num', end_value=float(rule['max']['value']), end_color=color(rule['max']['color'])))

    if data.get('filter'):
        ws.auto_filter = AutoFilter(ref=data['filter']['range'])

    # Sheets grows wrapped rows automatically; Excel needs an explicit height.
    for key, cell in cells.items():
        if not cell.get('wrap') or str(cell['row']) in data['rowHeights']:
            continue
        chars_per_line = max(1, int(data['colWidths'].get(str(cell['col']), 100) / 6.5))
        lines = sum(max(1, -(-len(part) // chars_per_line)) for part in str(shown.get(key) or '').split('\n'))
        if lines > 1:
            ws.row_dimensions[cell['row']].height = max(ws.row_dimensions[cell['row']].height or 0, 13.5 * lines)

    last_row = max(c['row'] for c in cells.values() if c.get('value') not in (None, ''))
    last_row = max([last_row] + [ch['position']['row'] + 16 for ch in data['charts']])
    last_col = max(int(c) for c in data['colWidths'])
    ws.print_area = f'A1:{get_column_letter(last_col)}{last_row}'
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    for chart in data['charts']:
        add_chart(ws, chart)

    if data.get('tabColor'):
        ws.sheet_properties.tabColor = color(data['tabColor'])


def add_chart(ws, spec):
    ranges = spec['ranges']
    opts = spec['options']
    kind = spec['type']
    if kind == 'PIE':
        ch = PieChart()
    elif kind == 'LINE':
        ch = LineChart()
    else:
        ch = BarChart()
        ch.type = 'bar' if kind == 'BAR' else 'col'
        if opts.get('isStacked'):
            ch.grouping = 'stacked'
            ch.overlap = 100
    ch.title = opts.get('title')
    ch.width, ch.height = opts.get('width', 600) / 37.8, opts.get('height', 300) / 37.8
    ch.legend.position = 'b'
    if opts.get('legend', {}).get('position') == 'none':
        ch.legend = None

    if len(ranges) == 1:
        c1, r1, c2, r2 = range_boundaries(ranges[0]['a1'])
        ch.add_data(Reference(ws, min_col=c1 + 1, max_col=c2, min_row=r1, max_row=r2), titles_from_data=True)
        ch.set_categories(Reference(ws, min_col=c1, min_row=r1 + 1, max_row=r2))
    else:  # one column per range: first = categories
        c1, r1, c2, r2 = range_boundaries(ranges[0]['a1'])
        for rng in ranges[1:]:
            cc1, rr1, cc2, rr2 = range_boundaries(rng['a1'])
            ch.add_data(Reference(ws, min_col=cc1, min_row=rr1, max_row=rr2), titles_from_data=True)
        ch.set_categories(Reference(ws, min_col=c1, min_row=r1 + 1, max_row=r2))

    colors = opts.get('colors')
    series_opts = opts.get('series', {})
    for i, s in enumerate(ch.series):
        col = series_opts.get(str(i), {}).get('color') or (colors[i] if colors and i < len(colors) else None)
        if col == 'transparent':
            s.graphicalProperties.noFill = True
            s.graphicalProperties.line.noFill = True
        elif col and kind != 'PIE':
            if kind == 'LINE':
                s.graphicalProperties.line.solidFill = color(col)
                s.graphicalProperties.line.width = 28000
                s.marker.symbol = 'circle'
                s.marker.graphicalProperties.solidFill = color(col)
            else:
                s.graphicalProperties.solidFill = color(col)
    if kind == 'PIE' and colors:
        from openpyxl.chart.series import DataPoint
        for i, col in enumerate(colors):
            pt = DataPoint(idx=i)
            pt.graphicalProperties.solidFill = color(col)
            ch.series[0].dPt.append(pt)
    if kind != 'PIE':
        axis = ch.y_axis  # openpyxl: y_axis is always the value axis, also for horizontal bars
        axis_opts = opts.get('hAxis' if kind == 'BAR' else 'vAxis', {})
        axis.number_format = '0%' if axis_opts.get('format') == 'percent' else '$#,##0.0,,"M"'
        view = axis_opts.get('viewWindow', {})
        lo = axis_opts.get('minValue', view.get('min'))
        hi = axis_opts.get('maxValue', view.get('max'))
        if lo is not None:
            axis.scaling.min = lo
        if hi is not None:
            axis.scaling.max = hi
        ch.y_axis.delete = False
        ch.x_axis.delete = False
    pos = spec['position']
    ws.add_chart(ch, f'{get_column_letter(pos["col"])}{pos["row"]}')


if __name__ == '__main__':
    main(sys.argv[1], sys.argv[2])
